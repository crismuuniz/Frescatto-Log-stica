import os
import io
import re
import pandas as pd
import reportlab
import pytesseract
import sqlite3

from flask import Flask, request, jsonify, render_template, send_file
from flask_sqlalchemy import SQLAlchemy
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from PIL import Image, ImageEnhance, ImageOps

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql+pymysql://usuario:senha@host_do_mysql/nome_do_banco'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

os.environ['TESSDATA_PREFIX'] = '/usr/share/tesseract-ocr/4.00/tessdata'
# Forçar o caminho do Tesseract no servidor Render
if os.path.exists('/usr/bin/tesseract'):
    pytesseract.pytesseract.tesseract_cmd = '/usr/bin/tesseract'

app = Flask(__name__)

# Se estiver rodando localmente no Windows e precisar apontar o Tesseract, mude aqui:
# pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
##TARIFAS REGRA DE NEGÓCIO (PAGAMENTO POR KM)
# ==================================

# ==================================
# TABELA DE TARIFAS E REGRA DE NEGÓCIO (PAGAMENTO POR KM)
# ==================================

def calcular_frete_por_veiculo(veiculo, rota, km=0, pedagio=0, diaria=0):
    veiculo = str(veiculo).strip().lower()
    rota = str(rota).strip().lower()

    tabela_fretes = {
        "fiorino": {
            "caravelas": 860.71,
            "alcobaça": 860.71,
            "prado": 860.71,
            "teixeira": 860.71,
            "cumuruxatiba": 860.71,
            "corumbau": 1114.67,
            "itamaraju": 1114.67,
            "mucuri": 1051.91,
            "nova viçosa": 1051.91,
            "nanuque": 1348.69,
            "montanha": 1348.69,
            "porto seguro": 1348.69,
            "cabrália": 1348.69,
            "cabralia": 1348.69,
            "arraial": 1348.69,
            "trancoso": 1348.69,
            "caraíva": 1348.69,
            "caraiva": 1348.69,
            "linhares": 1348.69,
            "são mateus": 1348.69,
            "sao mateus": 1348.69,
            "ilhéus": 3000.00,
            "ilheus": 3000.00,
            "valença": 4000.00,
            "valenca": 4000.00,
            "morro de são paulo": 4000.00,
            "morro de sao paulo": 4000.00,
            "itacaré": 4000.00,
            "itacare": 4000.00,
            "salvador": 4000.00
        },

        "bongo": {
            "caravelas": 1121.04,
            "alcobaça": 1121.04,
            "prado": 1121.04,
            "teixeira": 1121.04,
            "cumuruxatiba": 1121.04,
            "corumbau": 1369.47,
            "itamaraju": 1369.47,
            "mucuri": 1275.75,
            "nova viçosa": 1275.75,
            "nanuque": 1651.24,
            "montanha": 1651.24,
            "porto seguro": 1651.24,
            "cabrália": 1651.24,
            "cabralia": 1651.24,
            "arraial": 1651.24,
            "trancoso": 1651.24,
            "caraíva": 1651.24,
            "caraiva": 1651.24,
            "linhares": 1651.24,
            "são mateus": 1651.24,
            "sao mateus": 1651.24,
            "ilhéus": 5500.00,
            "ilheus": 5500.00,
            "valença": 7000.00,
            "valenca": 7000.00,
            "morro de são paulo": 7000.00,
            "morro de sao paulo": 7000.00,
            "itacaré": 7000.00,
            "itacare": 7000.00,
            "salvador": 7000.00
        },

        "hr": {
            "caravelas": 1121.04,
            "alcobaça": 1121.04,
            "prado": 1121.04,
            "teixeira": 1121.04,
            "cumuruxatiba": 1121.04,
            "corumbau": 1369.47,
            "itamaraju": 1369.47,
            "mucuri": 1275.75,
            "nova viçosa": 1275.75,
            "nanuque": 1651.24,
            "montanha": 1651.24,
            "porto seguro": 1651.24,
            "cabrália": 1651.24,
            "cabralia": 1651.24,
            "arraial": 1651.24,
            "trancoso": 1651.24,
            "caraíva": 1651.24,
            "caraiva": 1651.24,
            "linhares": 1651.24,
            "são mateus": 1651.24,
            "sao mateus": 1651.24,
            "ilhéus": 5500.00,
            "ilheus": 5500.00,
            "valença": 7000.00,
            "valenca": 7000.00,
            "morro de são paulo": 7000.00,
            "morro de sao paulo": 7000.00,
            "itacaré": 7000.00,
            "itacare": 7000.00,
            "salvador": 7000.00
        },

        "3/4": {
            "caravelas": 1296.37,
            "alcobaça": 1296.37,
            "prado": 1296.37,
            "teixeira": 1296.37,
            "cumuruxatiba": 1296.37,
            "corumbau": 1440.89,
            "itamaraju": 1440.89,
            "mucuri": 1474.32,
            "nova viçosa": 1474.32,
            "nanuque": 2119.91,
            "montanha": 2119.91,
            "porto seguro": 2119.91,
            "cabrália": 2119.91,
            "cabralia": 2119.91,
            "arraial": 2119.91,
            "trancoso": 2119.91,
            "caraíva": 2119.91,
            "caraiva": 2119.91,
            "linhares": 2119.91,
            "são mateus": 2119.91,
            "sao mateus": 2119.91,
            "salvador": 8800.00
        }
    }

    valor_base = 0

    if veiculo in tabela_fretes:
        for cidade, valor in tabela_fretes[veiculo].items():
            if cidade in rota:
                valor_base = valor
                break

    return round(valor_base + float(pedagio) + float(diaria), 2)
    


# ==================================
# SQLITE CONFIGURAÇÃO (ESTRUTURA PLANILHA FRESCATTO)
# ==================================

def get_db():
    # Corrigido: adicionada a conexão ao banco que estava faltando
    conn = sqlite3.connect("/tmp/database.db")
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    cursor = conn.cursor()

    # TABELA UNIFICADA: Estrutura real da planilha de rotas
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS rotas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        carga TEXT,
        data_carga TEXT,        -- Formato ISO: AAAA-MM-DD
        motorista TEXT,
        placa TEXT,
        veiculo TEXT,
        codigo_roteiro TEXT,    -- CÓD. ROT
        descricao_rota TEXT,    -- DESC. ROTA
        valor_coleta REAL,      -- VALOR COLET
        quantidade_entregas INTEGER, -- ENTREGAS
        peso REAL,
        volume REAL,
        valor_carga REAL,
        km REAL,
        pedagio REAL,
        diaria REAL,
        valor_frete REAL        -- VALOR FRETE (Faturamento calculado)
    )
    """)

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS canhotos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nota_fiscal TEXT,
        cliente TEXT,
        carga TEXT,
        data_entrega TEXT,
        status TEXT
    )
    """)

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS devolucoes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nota_fiscal TEXT,
        cliente TEXT,
        motivo TEXT,
        data TEXT,
        status TEXT
    )
    """)

    conn.commit()
    conn.close()


# Executa a inicialização do banco
init_db()
# ==================================


# ROTAS DE NAVEGAÇÃO (PÁGINAS)
# ==================================

@app.route("/")
def home():
    return render_template("index.html")

@app.route("/canhotos")
def canhotos_page():
    return render_template("canhotos.html")


@app.route("/fretes")
def fretes_page():
   return render_template("fretes.html")

@app.route("/roteiros")
def roteiros_page():
   return render_template("roteiros.html")

@app.route("/devolucoes")
def devolucoes_page():
   return render_template("devolucoes.html")

@app.route("/relatorios")
def relatorios_page():
    conn = get_db()
    rotas = conn.execute("SELECT * FROM rotas").fetchall()
    canhotos = conn.execute("SELECT * FROM canhotos").fetchall()
    devolucoes = conn.execute("SELECT * FROM devolucoes").fetchall()
    conn.close()

    return render_template(
        "relatorios.html",
        data="Relatório geral",
        rotas=rotas,
        canhotos=canhotos,
        devolucoes=devolucoes
    )

# ==================================
# API: MÓDULO UNIFICADO DE ROTAS E FRETES (COM CÁLCULO DE KM)
# ==================================

# GET: Retorna as rotas
@app.route("/api/roteiros", methods=["POST"])
def adicionar_roteiro():
    dados = request.json
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO rotas (motorista, veiculo, codigo_roteiro, data_carga, km, pedagio, diaria, 
                           carga, descricao_rota, quantidade_entregas, peso, volume) 
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (dados['motorista'], dados['veiculo'], dados['codigo_roteiro'], dados['data_carga'], 
          dados['km'], dados['pedagio'], dados['diaria'], dados['carga'], dados['descricao_rota'], 
          dados['quantidade_entregas'], dados['peso'], dados['volume']))
    conn.commit()
    return jsonify({"status": "sucesso"})

# PUT: Atualiza as informações recalculando o valor do frete
@app.route("/api/roteiros/<int:id>", methods=["PUT"])
@app.route("/api/fretes/<int:id>", methods=["PUT"])
def update_rota(id):
    dados = request.json
    conn = get_db()

    veiculo = dados.get("veiculo") or dados.get("tipo_veiculo") or "Padrão"
    rota = (dados.get("descricao_rota") or dados.get("rota") or "")
    km = float(dados.get("km") or 0)
    pedagio = float(dados.get("pedagio") or 0)
    diaria = float(dados.get("diaria") or 0)

    valor_final_frete = calcular_frete_por_veiculo(veiculo, rota, km, pedagio, diaria)

    conn.execute("""
        UPDATE rotas SET
            carga = ?, data_carga = ?, motorista = ?, placa = ?, veiculo = ?,
            codigo_roteiro = ?, descricao_rota = ?, valor_coleta = ?,
            quantidade_entregas = ?, peso = ?, volume = ?, valor_carga = ?,
            km = ?, pedagio = ?, diaria = ?, valor_frete = ?
        WHERE id = ?
    """, (
        dados.get("carga") or dados.get("rota"),
        dados.get("data_carga") or dados.get("data"),
        dados.get("motorista"), dados.get("placa"), veiculo,
        dados.get("codigo_roteiro") or dados.get("romaneio"),
        dados.get("descricao_rota"), dados.get("valor_coleta", 0),
        dados.get("quantidade_entregas", 0), dados.get("peso", 0),
        dados.get("volume", 0), dados.get("valor_carga", 0),
        km, pedagio, diaria, valor_final_frete, id
    ))

    conn.commit()
    conn.close()

    # O return deve estar recuado (dentro da função)
    return jsonify({"status": "ok", "valor_calculado": valor_final_frete})
# ==================================
# API: CANHOTOS
# ==================================

@app.route("/api/canhotos", methods=["GET"])
def get_canhotos():
    conn = get_db()
    rows = conn.execute("SELECT * FROM canhotos").fetchall()
    conn.close()

    return jsonify([dict(r) for r in rows])


@app.route("/api/canhotos", methods=["POST"])
def add_canhoto():
    dados = request.json
    conn = get_db()

    conn.execute("""
        INSERT INTO canhotos (
            nota_fiscal,
            cliente,
            carga,
            data_entrega,
            status
        )
        VALUES (?, ?, ?, ?, ?)
    """, (
        dados.get("nota_fiscal"),
        dados.get("cliente"),
        dados.get("carga"),
        dados.get("data_entrega"),
        dados.get("status", "Pendente")
    ))

    conn.commit()
    conn.close()

    return jsonify({"status": "ok"}), 201

@app.route("/api/canhotos/<int:id>", methods=["DELETE"])
def delete_canhoto(id):
    conn = get_db()
    conn.execute("DELETE FROM canhotos WHERE id = ?", (id,))
    conn.commit()
    conn.close()
    return jsonify({"status": "ok"})


# ==================================
# API: DEVOLUÇÕES (CRUD COMPLETO)
# ==================================

@app.route("/api/devolucoes", methods=["GET"])
def get_devolucoes():
    conn = get_db()
    rows = conn.execute("SELECT * FROM devolucoes").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/devolucoes", methods=["POST"])
def add_devolucao():
    dados = request.json
    conn = get_db()

    conn.execute("""
        INSERT INTO devolucoes (
            nota_fiscal,
            cliente,
            motivo,
            data,
            status
        )
        VALUES (?, ?, ?, ?, ?)
    """, (
        dados.get("nota_fiscal"),
        dados.get("cliente"),
        dados.get("motivo"),
        dados.get("data"),
        dados.get("status", "Pendente")
    ))

    conn.commit()
    conn.close()

    return jsonify({"status": "ok"}), 201


@app.route("/api/devolucoes/<int:id>", methods=["PUT"])
def update_devolucao(id):
    dados = request.json
    conn = get_db()

    conn.execute("""
        UPDATE devolucoes
        SET
            nota_fiscal = ?,
            cliente = ?,
            motivo = ?,
            data = ?,
            status = ?
        WHERE id = ?
    """, (
        dados.get("nota_fiscal"),
        dados.get("cliente"),
        dados.get("motivo"),
        dados.get("data"),
        dados.get("status"),
        id
    ))

    conn.commit()
    conn.close()

    return jsonify({"status": "ok"})


@app.route("/api/devolucoes/<int:id>", methods=["DELETE"])
def delete_devolucao(id):
    conn = get_db()
    conn.execute("DELETE FROM devolucoes WHERE id = ?", (id,))
    conn.commit()
    conn.close()

    return jsonify({"status": "ok"})


# ==================================
# INTEGRAÇÃO: PROCESSAMENTO DE IMAGEM (OCR)
# ==================================

@app.route("/api/processar-imagem", methods=["POST"])
def processar_imagem():
    if "imagem" not in request.files:
        return jsonify({"erro": "Nenhuma imagem enviada"}), 400

    arquivo_imagem = request.files["imagem"]

    if arquivo_imagem.filename == "":
        return jsonify({"erro": "Arquivo inválido"}), 400

    try:
        imagem = Image.open(arquivo_imagem)
        texto_extraido = pytesseract.image_to_string(imagem, lang="por")

        padrao_nf = re.search(
            r"(?:nota|nf|nº|numero|número)[:.\s\-]+(\d+)",
            texto_extraido,
            re.IGNORECASE
        )
        nota_fiscal = padrao_nf.group(1) if padrao_nf else ""

        padrao_valor = re.search(
            r"(?:total|valor|r\$)[:\s\-]+([\d.,]+)",
            texto_extraido,
            re.IGNORECASE
        )
        valor_total = padrao_valor.group(1) if padrao_valor else ""

        padrao_cliente = re.search(
            r"(?:cliente|destinatario|razão social)[:\s\-]+([A-Za-z0-9\s]+)",
            texto_extraido,
            re.IGNORECASE
        )
        cliente = padrao_cliente.group(1).strip() if padrao_cliente else ""

        return jsonify({
            "status": "sucesso",
            "dados_sugeridos": {
                "nota_fiscal": nota_fiscal,
                "cliente": cliente,
                "valor_total": valor_total
            },
            "texto_completo_identificado": texto_extraido
        })

    except Exception as e:
        return jsonify({
            "erro": f"Falha no processamento: {str(e)}"
        }), 500


# ==================================
# EMISSÃO DE PDF ADAPTADO (NOVA TABELA)
# ==================================

@app.route("/relatorio/intervalo/pdf")
def relatorio_pdf():
    inicio = request.args.get("inicio") or "2000-01-01"
    fim = request.args.get("fim") or "2100-12-31"

    conn = get_db()

    rotas = conn.execute(
        "SELECT * FROM rotas WHERE data_carga BETWEEN ? AND ?",
        (inicio, fim)
    ).fetchall()

    canhotos = conn.execute(
        "SELECT * FROM canhotos WHERE data_entrega BETWEEN ? AND ?",
        (inicio, fim)
    ).fetchall()

    devolucoes = conn.execute(
        "SELECT * FROM devolucoes WHERE data BETWEEN ? AND ?",
        (inicio, fim)
    ).fetchall()

    resumo_motoristas = conn.execute("""
        SELECT
            motorista,
            COUNT(*) AS viagens,
            SUM(valor_frete) AS total
        FROM rotas
        WHERE data_carga BETWEEN ? AND ?
        GROUP BY motorista
        ORDER BY total DESC
    """, (inicio, fim)).fetchall()

    conn.close()

    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)

    y = 800

    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(
        50,
        y,
        f"Relatório Logístico {inicio} até {fim}"
    )

    y -= 40

    def bloco(titulo, dados, texto):
        nonlocal y

        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(50, y, titulo)

        y -= 20

        pdf.setFont("Helvetica", 10)

        for d in dados:
            pdf.drawString(50, y, texto(d))
            y -= 15

            if y < 50:
                pdf.showPage()
                y = 800

        y -= 10

    bloco(
        "MOVIMENTAÇÃO DE ROTAS E FRETES",
        rotas,
        lambda r: (
            f"Carga: {r['carga']} | "
            f"{r['motorista']} | "
            f"{r['descricao_rota']} | "
            f"Frete: R$ {r['valor_frete']:.2f}"
        )
    )

    bloco(
        "CANHOTOS DE NOTAS",
        canhotos,
        lambda c: (
            f"NF: {c['nota_fiscal']} | "
            f"Carga: {c['carga']} | "
            f"Status: {c['status']}"
        )
    )

    bloco(
        "DEVOLUÇÕES CONSTATADAS",
        devolucoes,
        lambda d: (
            f"NF: {d['nota_fiscal']} | "
            f"Motivo: {d['motivo']} | "
            f"{d['status']}"
        )
    )

    if y < 150:
        pdf.showPage()
        y = 800

    pdf.setFont("Helvetica-Bold", 13)
    pdf.drawString(
        50,
        y,
        "FECHAMENTO FINANCEIRO POR MOTORISTA"
    )

    y -= 25

    total_geral = 0

    pdf.setFont("Helvetica", 10)

    for r in resumo_motoristas:
        total_geral += r["total"] if r["total"] else 0

        pdf.drawString(
            50,
            y,
            f"{r['motorista']} | "
            f"Viagens: {r['viagens']} | "
            f"Total Faturado: R$ {r['total']:.2f}"
        )

        y -= 15

    y -= 15

    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(
        50,
        y,
        f"TOTAL GERAL DO PERÍODO: R$ {total_geral:.2f}"
    )

    pdf.save()

    buffer.seek(0)

    return send_file(
        buffer,
        download_name=f"relatorio_{inicio}_{fim}.pdf",
        as_attachment=True
    )

@app.route("/api/exportar-excel")
def exportar_excel():
    try:
        inicio = request.args.get("inicio") or "2000-01-01"
        fim = request.args.get("fim") or "2100-12-31"

        conn = get_db()
        query = "SELECT * FROM rotas WHERE data_carga BETWEEN ? AND ?"
        df = pd.read_sql_query(query, conn, params=(inicio, fim))
        conn.close()

        if df.empty:
            return "Nenhum dado encontrado.", 404

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 1. Exportar dados começando na linha 3
            df.to_excel(writer, index=False, sheet_name='Fretes', startrow=2)
            
            workbook = writer.book
            worksheet = writer.sheets['Fretes']
            
            # 2. Título Dinâmico (Mescla baseado na quantidade de colunas)
            num_cols = len(df.columns)
            ultima_coluna_letra = get_column_letter(num_cols)
            intervalo_titulo = f'A1:{ultima_coluna_letra}1'
            
            worksheet.merge_cells(intervalo_titulo)
            titulo = worksheet['A1']
            titulo.value = "TABELA DE FRETES MOTORISTAS TERCEIRIZADOS"
            titulo.font = Font(size=16, bold=True, color="FFFFFF")
            titulo.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            titulo.alignment = Alignment(horizontal="center")

            # 3. Formatação dos Cabeçalhos (Linha 3)
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2F3542", end_color="2F3542", fill_type="solid")
            for cell in worksheet[3]:
                cell.font = header_font
                cell.fill = header_fill

            # 4. Formatação de Moeda e Ajuste de Largura
        for idx, col in enumerate(df.columns):
            col_letter = get_column_letter(idx + 1)
            
            # Ajuste de largura automático
            max_len = max(df[col].astype(str).map(len).max(), len(col)) + 4
            worksheet.column_dimensions[col_letter].width = max_len

            # Formato de moeda para colunas financeiras
            if any(termo in col.lower() for termo in ['valor', 'frete', 'diaria', 'pedagio']):
                # iter_rows garante que sempre tenhamos um iterável, evitando o erro de 'tuple'
                for row in worksheet.iter_rows(min_row=4, min_col=idx+1, max_col=idx+1, max_row=len(df)+3):
                    for cell in row:
                        cell.number_format = 'R$ #,##0.00'

        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            download_name='Relatorio_Financeiro_Frescatto.xlsx',
            as_attachment=True
        )
    except Exception as e:
        # Se der erro, o navegador mostrará exatamente o que aconteceu
        return f"Erro ao gerar o arquivo: {str(e)}", 500

@app.route("/api/exportar-devolucoes")
def exportar_devolucoes_excel():
    try:
        inicio = request.args.get("inicio") or "2000-01-01"
        fim = request.args.get("fim") or "2100-12-31"

        conn = get_db()
        query = "SELECT * FROM devolucoes WHERE data BETWEEN ? AND ?"
        df = pd.read_sql_query(query, conn, params=(inicio, fim))
        conn.close()

        if df.empty:
            return "Nenhuma devolução encontrada no período.", 404

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Devolucoes', startrow=2)
            
            worksheet = writer.sheets['Devolucoes']
            
            # Título
            num_cols = len(df.columns)
            col_letter = get_column_letter(num_cols)
            worksheet.merge_cells(f'A1:{col_letter}1')
            titulo = worksheet['A1']
            titulo.value = "RELATÓRIO DE DEVOLUÇÕES"
            titulo.font = Font(size=16, bold=True, color="FFFFFF")
            titulo.fill = PatternFill(start_color="2F3542", end_color="2F3542", fill_type="solid")
            titulo.alignment = Alignment(horizontal="center")

            # Formatação de cabeçalhos
            for cell in worksheet[3]:
                cell.font = Font(bold=True)

            # Ajuste de largura
            for idx, col in enumerate(df.columns):
                col_let = get_column_letter(idx + 1)
                max_len = max(df[col].astype(str).map(len).max(), len(col)) + 4
                worksheet.column_dimensions[col_let].width = max_len

        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            download_name='Relatorio_Devolucoes.xlsx',
            as_attachment=True
        )
    except Exception as e:
        return f"Erro: {str(e)}", 500

@app.route("/api/exportar-roteiros")
def exportar_roteiros_excel():
    try:
        inicio = request.args.get("inicio") or "2000-01-01"
        fim = request.args.get("fim") or "2100-12-31"

        conn = get_db()
        # Filtra os roteiros pelo período
        query = "SELECT * FROM rotas WHERE data_carga BETWEEN ? AND ?"
        df = pd.read_sql_query(query, conn, params=(inicio, fim))
        conn.close()

        if df.empty:
            return "Nenhum roteiro encontrado no período.", 404

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Roteiros', startrow=2)
            worksheet = writer.sheets['Roteiros']
            
            # Título e Estilo (mesmo padrão dos outros)
            num_cols = len(df.columns)
            col_letter = get_column_letter(num_cols)
            worksheet.merge_cells(f'A1:{col_letter}1')
            worksheet['A1'] = "RELATÓRIO DE ROTEIROS"
            worksheet['A1'].font = Font(size=16, bold=True, color="FFFFFF")
            worksheet['A1'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            worksheet['A1'].alignment = Alignment(horizontal="center")

            # Formatação de cabeçalho
            for cell in worksheet[3]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="2F3542", end_color="2F3542", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)

            # Ajuste de largura e formatação (Corrigida a indentação)
            for idx, col in enumerate(df.columns):
                col_let = get_column_letter(idx + 1)
                max_len = max(df[col].astype(str).map(len).max(), len(col)) + 4
                worksheet.column_dimensions[col_let].width = max_len
                
                # Formato de moeda se for coluna financeira
                if any(termo in col.lower() for termo in ['valor', 'frete', 'diaria', 'pedagio']):
                    for row in worksheet.iter_rows(min_row=4, min_col=idx+1, max_col=idx+1, max_row=len(df)+3):
                        for cell in row:
                            cell.number_format = 'R$ #,##0.00'

        output.seek(0)
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         download_name='Relatorio_Roteiros.xlsx', as_attachment=True)
    except Exception as e:
        return f"Erro: {str(e)}", 500

# Rota simples para receber o texto extraído do OCR
@app.route("/api/salvar-dados", methods=["POST"])
def salvar_dados():
    dados = request.json
    texto = dados.get("texto_ocr", "")
    
    # Aqui você pode aplicar sua lógica de Regex para extrair nota/valor/cliente
    # e salvar no seu SQLite (database.db)
    
    print("Texto recebido do frontend:", texto)
    return jsonify({"status": "sucesso", "mensagem": "Dados gravados no banco!"}), 200
        
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 3000)))